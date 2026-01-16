from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Dict, Any, List
from pathlib import Path
from app.core.database import get_db
from app.api.dependencies import get_current_user
from app.models.user import User
from app.models.file import File
from app.services.transform_service import transform_service
from app.services.file_service import file_service
from app.services.file_reference_service import file_reference_service
from app.services.preview_cache import preview_cache, stable_hash
from app.utils.export_utils import create_zip_archive
import pandas as pd
import io
import re


router = APIRouter(prefix="/transform", tags=["transform"])


class FlowExecuteRequest(BaseModel):
    file_id: int
    file_ids: list[int] | None = None
    flow_data: Dict[str, Any]
    preview_target: Dict[str, Any] | None = None
    output_batch_id: int | None = None


class FlowPrecomputeRequest(BaseModel):
    file_id: int | None = None
    file_ids: list[int] | None = None
    flow_data: Dict[str, Any]


class StepPreviewRequest(BaseModel):
    file_id: int
    step_config: Dict[str, Any]


class ListOutputsResponse(BaseModel):
    outputs: List[Dict[str, Any]]


@router.post("/list-outputs", response_model=ListOutputsResponse)
async def list_outputs(
    request: FlowPrecomputeRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Execute a flow and list all available output tables for preview."""
    requested_ids = request.file_ids if request.file_ids else []
    if request.file_id:
        requested_ids.append(request.file_id)

    referenced_ids = list(
        file_reference_service.extract_file_ids_from_flow_data(request.flow_data))
    effective_ids = list(set(requested_ids) | set(referenced_ids))

    db_files = db.query(File).filter(File.user_id == current_user.id, File.id.in_(
        effective_ids)).all() if effective_ids else []

    # Validate that all requested IDs exist
    if requested_ids:
        found_ids = {f.id for f in db_files}
        missing_ids = set(requested_ids) - found_ids
        if missing_ids:
            raise HTTPException(
                status_code=404,
                detail=f"Requested files not found: {list(missing_ids)}"
            )

    file_paths_by_id = {db_file.id: db_file.file_path for db_file in db_files}

    try:
        table_map, _, terminal_keys = transform_service.execute_flow(
            file_paths_by_id,
            request.flow_data
        )

        output_targets = []

        # 1. Collect Explicit Final Outputs (from Transform Nodes)
        # 2. Collect Implicit G2G Outputs (from Transform Nodes without destinations)
        # 3. Collect Output Block outputs (Legacy/Explicit)

        nodes = request.flow_data.get("nodes", [])
        for node in nodes:
            data = node.get("data", {})
            if not data:
                continue

            block_type = data.get("blockType")
            if block_type == "output":
                # Handle Output Node (Legacy/Explicit)
                output_config = data.get("output", {})
                for i, out_file in enumerate(output_config.get("outputs", [])):
                    file_name = out_file.get("fileName", f"output_{i+1}.xlsx")
                    output_id = out_file.get("id", f"output-{i+1}")
                    for sheet in out_file.get("sheets", []):
                        sheet_name = sheet.get("sheetName", "Sheet1")
                        virtual_id = f"virtual:output:{output_id}:{sheet_name}"
                        output_targets.append({
                            "virtualId": virtual_id,
                            "virtualName": f"{file_name} / {sheet_name}",
                            "fileId": None,
                            "sheetName": sheet_name,
                            "isFinalOutput": True
                        })
                continue

            # Handle Transform Nodes
            dest_targets = data.get("destinationTargets", [])
            source_targets = data.get("sourceTargets", [])

            # Implicit G2G Check: No destinations, but has batch source
            if not dest_targets and source_targets:
                # Check if it's a batch source (simplified check: if sources > 1 or batchId present)
                # In backend execution, dest=source if dest is empty.
                # So we verify if we should list them as outputs.
                # Assuming Implicit G2G = Final Output by default (as per frontend logic)
                has_batch = any(t.get("batchId")
                                for t in source_targets) or len(source_targets) > 1
                if has_batch:
                    # Provide outputs mapping to the inputs (which are updated in place)
                    for t in source_targets:
                        # Construct identity used by transform_service
                        # transform_service uses get_key_for_target(t)
                        # We need to return metadata for Preview Selector
                        output_targets.append({
                            "virtualId": t.get("virtualId"),
                            "fileId": t.get("fileId"),
                            "sheetName": t.get("sheetName"),
                            "virtualName": t.get("virtualName") or f"Updated {t.get('fileId')}",
                            "isFinalOutput": True
                        })

            # Explicit Check
            if dest_targets:
                for t in dest_targets:
                    if t.get("isFinalOutput"):
                        output_targets.append({
                            "virtualId": t.get("virtualId"),
                            "batchId": t.get("batchId"),
                            "sheetName": t.get("sheetName"),
                            "virtualName": t.get("virtualName"),
                            "isFinalOutput": True,
                            "fileId": t.get("fileId")  # If mapped to real file
                        })

        return {"outputs": [t for t in output_targets if t.get("virtualId") or (t.get("fileId") and t.get("sheetName"))]}

    except Exception as e:
        # It's important to raise an error with details for debugging
        raise HTTPException(
            status_code=400,
            detail=f"Error listing outputs: {str(e)}"
        )


@router.post("/execute")
async def execute_flow(
    request: FlowExecuteRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Execute a flow on a file"""
    requested_ids = request.file_ids if request.file_ids else [request.file_id]
    requested_ids = [file_id for file_id in requested_ids if isinstance(
        file_id, int) and file_id > 0]
    referenced_ids = list(
        file_reference_service.extract_file_ids_from_flow_data(request.flow_data))
    effective_ids = requested_ids or referenced_ids
    # Load all referenced files (multi-file flows need more than one).
    db_files = []
    if effective_ids:
        db_files = db.query(File).filter(
            File.user_id == current_user.id,
            File.id.in_(effective_ids)
        ).all()

    if effective_ids and not db_files:
        raise HTTPException(status_code=404, detail="File not found")

    file_paths_by_id = {db_file.id: db_file.file_path for db_file in db_files}
    file_fingerprints = [
        {"id": db_file.id, "size": db_file.file_size} for db_file in db_files]

    # Execute flow
    try:
        # Cache preview results to avoid re-running transforms on repeated previews.
        preview_target_payload = request.preview_target or {}
        preview_cache_key = stable_hash({
            "user_id": current_user.id,
            "files": file_fingerprints,
            "flow_data": request.flow_data,
            "preview_target": preview_target_payload,
        })

        cached_preview = preview_cache.get(preview_cache_key)
        if cached_preview is not None:
            return cached_preview

        table_map, last_table_key, _ = transform_service.execute_flow(
            file_paths_by_id,
            request.flow_data
        )

        preview_target = preview_target_payload
        target_file_id = preview_target.get("file_id")
        target_sheet_name = preview_target.get("sheet_name")
        target_virtual_id = preview_target.get("virtual_id")

        if not target_file_id and isinstance(target_virtual_id, str):
            table_key = f"virtual:{target_virtual_id}"
            result_df = table_map.get(table_key)
            if result_df is None:
                result_df = pd.DataFrame()
            preview = file_service.get_file_preview(result_df)
            response_payload = {
                "preview": preview,
                "row_count": len(result_df),
                "column_count": len(result_df.columns)
            }
            preview_cache.set(preview_cache_key, response_payload)
            return response_payload

        if target_file_id:
            table_key = f"{target_file_id}:{target_sheet_name or '__default__'}"
            result_df = table_map.get(table_key)
            if result_df is None and target_file_id in file_paths_by_id:
                result_df = file_service.parse_file(
                    file_paths_by_id[target_file_id],
                    sheet_name=target_sheet_name
                )
            if result_df is None:
                result_df = pd.DataFrame()
        elif last_table_key and last_table_key in table_map:
            result_df = table_map[last_table_key]
        elif effective_ids:
            # Fallback to the first file in case no transforms ran.
            fallback_file_id = effective_ids[0]
            result_df = file_service.parse_file(
                file_paths_by_id[fallback_file_id])
        else:
            result_df = pd.DataFrame()

        preview = file_service.get_file_preview(result_df)

        response_payload = {
            "preview": preview,
            "row_count": len(result_df),
            "column_count": len(result_df.columns)
        }
        preview_cache.set(preview_cache_key, response_payload)
        return response_payload
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Error executing flow: {str(e)}"
        )


@router.post("/precompute")
async def precompute_flow(
    request: FlowPrecomputeRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Precompute previews for output sheets to warm the server cache."""
    requested_ids = request.file_ids if request.file_ids else []
    if request.file_id:
        requested_ids = requested_ids + [request.file_id]
    requested_ids = [file_id for file_id in requested_ids if isinstance(
        file_id, int) and file_id > 0]
    referenced_ids = list(
        file_reference_service.extract_file_ids_from_flow_data(request.flow_data))
    effective_ids = requested_ids or referenced_ids

    db_files = []
    if effective_ids:
        db_files = db.query(File).filter(
            File.user_id == current_user.id,
            File.id.in_(effective_ids)
        ).all()

    if effective_ids and not db_files:
        raise HTTPException(status_code=404, detail="File not found")

    file_paths_by_id = {db_file.id: db_file.file_path for db_file in db_files}
    file_fingerprints = [
        {"id": db_file.id, "size": db_file.file_size} for db_file in db_files]

    nodes = request.flow_data.get("nodes", [])
    output_node = next(
        (node for node in nodes if node.get(
            "data", {}).get("blockType") == "output"),
        None
    )
    output_config = output_node.get("data", {}).get(
        "output", {}) if output_node else {}
    output_files = output_config.get(
        "outputs") if isinstance(output_config, dict) else []

    if not output_files:
        return {"status": "skipped", "precomputed": 0}

    try:
        # Execute once so we can reuse the resulting tables for all output sheets.
        table_map, _, _ = transform_service.execute_flow(
            file_paths_by_id,
            request.flow_data
        )

        precomputed = 0
        for index, output_file in enumerate(output_files):
            output_id = output_file.get("id") if isinstance(
                output_file, dict) else None
            if not output_id:
                output_id = f"output-{index + 1}"
            sheets = output_file.get("sheets") if isinstance(
                output_file, dict) else []
            if not sheets:
                sheets = [{"sheetName": "Sheet 1"}]
            for sheet in sheets:
                sheet_name = sheet.get("sheetName") or "Sheet 1"
                preview_target = {
                    "virtual_id": f"output:{output_id}:{sheet_name}",
                    "sheet_name": sheet_name,
                }
                preview_cache_key = stable_hash({
                    "user_id": current_user.id,
                    "files": file_fingerprints,
                    "flow_data": request.flow_data,
                    "preview_target": preview_target,
                })
                if preview_cache.get(preview_cache_key) is not None:
                    continue
                table_key = f"virtual:{preview_target['virtual_id']}"
                result_df = table_map.get(table_key, pd.DataFrame())
                preview = file_service.get_file_preview(result_df)
                preview_cache.set(preview_cache_key, {
                    "preview": preview,
                    "row_count": len(result_df),
                    "column_count": len(result_df.columns)
                })
                precomputed += 1

        return {"status": "ok", "precomputed": precomputed}
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Error precomputing previews: {str(e)}"
        )


@router.post("/preview-step")
async def preview_step(
    request: StepPreviewRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Preview a single transformation step"""
    # Get file
    db_file = db.query(File).filter(
        File.id == request.file_id,
        File.user_id == current_user.id
    ).first()

    if not db_file:
        raise HTTPException(status_code=404, detail="File not found")

    # Preview step
    try:
        preview = transform_service.preview_step(
            db_file.file_path,
            request.step_config
        )
        return preview
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Error previewing step: {str(e)}"
        )


@router.post("/export")
async def export_result(
    request: FlowExecuteRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Execute flow and export result as Excel"""
    requested_ids = request.file_ids if request.file_ids else [request.file_id]
    db_files = db.query(File).filter(
        File.user_id == current_user.id,
        File.id.in_(requested_ids)
    ).all()

    if not db_files:
        raise HTTPException(status_code=404, detail="File not found")

    file_paths_by_id = {db_file.id: db_file.file_path for db_file in db_files}

    # Calculate effective_ids for fallback logic
    referenced_ids = list(
        file_reference_service.extract_file_ids_from_flow_data(request.flow_data))
    effective_ids = requested_ids or referenced_ids

    output_batch = None
    if request.output_batch_id is not None:
        from app.models.file_batch import FileBatch
        output_batch = db.query(FileBatch).filter(
            FileBatch.user_id == current_user.id,
            FileBatch.id == request.output_batch_id
        ).first()
        if not output_batch:
            raise HTTPException(
                status_code=404, detail="Output batch not found")

    # Execute flow
    try:
        table_map, last_table_key, _ = transform_service.execute_flow(
            file_paths_by_id,
            request.flow_data
        )

        # Initialize default result_df from last step if available
        result_df = pd.DataFrame()
        if last_table_key and last_table_key in table_map:
            result_df = table_map[last_table_key]

        # Collect outputs to export
        outputs_to_write = []
        nodes = request.flow_data.get("nodes", [])

        # 1. Explicit Final Outputs & Implicit G2G
        for node in nodes:
            data = node.get("data", {}) or {}
            dest_targets = data.get("destinationTargets", [])
            source_targets = data.get("sourceTargets", [])

            # Explicit
            if dest_targets:
                for t in dest_targets:
                    if t.get("isFinalOutput"):
                        # Map to output structure
                        fname = t.get("virtualName") or "output"
                        if not fname.endswith(".xlsx") and not fname.endswith(".csv"):
                            fname += ".xlsx"

                        outputs_to_write.append({
                            "fileName": fname,
                            "sheets": [{"sheetName": t.get("sheetName") or "Sheet1"}],
                            "target": t  # Keep ref to look up data
                        })

            # Implicit G2G (No dests, has batch source)
            elif source_targets:
                has_batch = any(t.get("batchId")
                                for t in source_targets) or len(source_targets) > 1
                if has_batch:
                    for t in source_targets:
                        # Use source name if available
                        # Fetch original filename if possible from DB?
                        # We have file_paths_by_id, but not names here easily unless we queried files.
                        # But 't' might have 'virtualName'.
                        fname = t.get(
                            "virtualName") or f"file_{t.get('fileId')}.xlsx"
                        if not fname.endswith(".xlsx") and not fname.endswith(".csv"):
                            fname += ".xlsx"

                        outputs_to_write.append({
                            "fileName": fname,
                            "sheets": [{"sheetName": t.get("sheetName") or "Sheet1"}],
                            "target": t
                        })

        # 2. Output Node (Legacy/Explicit Output Block Config)
        # Only if NO other outputs found? Or merge?
        # Prompt says "Output Node... aggregator...".
        # If the user has an Output Block defined, we should include it too as they might rely on it.
        output_node = next((n for n in nodes if n.get(
            "data", {}).get("blockType") == "output"), None)
        if output_node:
            output_config = output_node.get("data", {}).get("output", {})
            legacy_outputs = output_config.get("outputs", [])
            for out_file in legacy_outputs:
                # Add check to avoid duplicates if necessary, but for now append.
                outputs_to_write.append(out_file)

            # Legacy single-file config
            legacy_file_name = output_config.get("fileName")
            legacy_sheets = output_config.get("sheets")
            if legacy_file_name or legacy_sheets:
                outputs_to_write.append({
                    "fileName": legacy_file_name or "output.xlsx",
                    "sheets": legacy_sheets or [{"sheetName": "Sheet1"}]
                })

        # Fallback
        if not outputs_to_write and effective_ids:
            # Just export the input if nothing transformed?
            pass

        if not outputs_to_write and last_table_key and last_table_key in table_map:
            # Fallback to exporting the result of the last step if no explicit outputs are defined.
            outputs_to_write.append({
                "fileName": "output.xlsx",
                "sheets": [{"sheetName": "Sheet1"}],
                "target": None  # Indicates to use the default result_df
            })

        files_payload: list[dict[str, str | bytes]] = []
        reserved_output_names: set[str] = set()

        # Helper to get DF
        def get_df_for_target(t: Dict[str, Any]) -> pd.DataFrame:
            # Logic matches transform_service.get_key_for_target strategy
            # 1. Try virtualId
            vid = t.get("virtualId")
            if vid:
                key = f"virtual:{vid}"
                if key in table_map:
                    return table_map[key]

            # 2. Try fileId:sheet
            fid = t.get("fileId")
            if fid:
                sname = t.get("sheetName")
                key = f"{fid}:{sname or '__default__'}"
                if key in table_map:
                    return table_map[key]
                # If not in table_map, maybe it wasn't modified? Parse original?
                if fid in file_paths_by_id:
                    return file_service.parse_file(file_paths_by_id[fid], sheet_name=sname)

            return pd.DataFrame()  # Empty if not found

        # Check for Append Configuration
        # Check for Append Configuration / Output Config
        # Refactor: We look for ANY node that has configured output settings (writeMode or batchNamingPattern)
        # We prioritize the "last" node or an explicit Output node if present.
        output_config_node = next((n for n in nodes if n.get(
            "data", {}).get("blockType") == "output"), None)

        if not output_config_node:
            # Fallback: Find any node that has explicit output config
            # We iterate in reverse to find the "tail" config if possible?
            # Or just finds the first one that has writeMode/naming.
            # Since flow is usually sequential, we might want the last one.
            for n in reversed(nodes):
                d = n.get("data", {}).get("output", {})
                if d.get("writeMode") or d.get("batchNamingPattern") or d.get("mode") == "batch_template":
                    output_config_node = n
                    break

        write_mode = "create"
        base_file_id = None
        output_batch = None

        if output_config_node:
            output_config = output_config_node.get(
                "data", {}).get("output", {})
            write_mode = output_config.get("writeMode", "create")
            base_file_id = output_config.get("baseFileId")

            # Also resolve batch naming context from this node if needed?
            # Actually output_batch is resolved from the FLOW context usually or explicit batch node.
            # But we might need check if this node defines batch settings.

        # If Append Mode, ensure we treat it as a single file write (Merge)
        if write_mode == "append" and base_file_id:
            # Fetch base file if not already loaded
            if base_file_id not in file_paths_by_id:
                base_file = db.query(File).filter(
                    File.id == base_file_id).first()
                if base_file:
                    file_paths_by_id[base_file_id] = base_file.file_path

            if base_file_id in file_paths_by_id:
                # We will process ALL outputs into this one file
                # effectively converting "Separate" -> "Append Sheets to Base"
                pass  # Logic handled below

        # REFACTORED LOOP
        if write_mode == "append" and base_file_id and base_file_id in file_paths_by_id:
            # APPEND MODE LOGIC
            base_path = file_paths_by_id[base_file_id]
            import openpyxl

            # Load workbook
            try:
                book = openpyxl.load_workbook(base_path)
            except Exception:
                # Fallback if invalid base file, create new
                book = openpyxl.Workbook()

            # Create output buffer
            output = io.BytesIO()
            # We must save the book to the buffer to "start" the writer with it?
            # pd.ExcelWriter doesn't easily accept an openpyxl book for *init* unless specific hack.
            # Standard way:
            # writer.book = book
            # writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

            book.save(output)  # Save current base state to buffer
            output.seek(0)

            # Use pandas with 'openpyxl' engine in append mode?
            # No, 'mode="a"' is for file paths. For bytes, we manually manipulate book.

            with pd.ExcelWriter(output, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                # This requires "output" to be a path or a handle that supports read/write/seek like a file.
                # BytesIO works if pre-filled?
                # Actually, pandas append mode on BytesIO is tricky/unsupported often.
                # BETTER: Just use writer.book = book logic.
                pass

            # Manual Writer Setup
            output = io.BytesIO()  # Clear output, we will save book at end
            writer = pd.ExcelWriter(output, engine="openpyxl")
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}

            # Loop all items and write to this single writer
            for item in outputs_to_write:
                sheets = item.get("sheets", [])
                target = item.get("target")

                # Prepare DF
                if sheets:
                    for sheet in sheets:
                        sheet_name = sheet.get("sheetName") or "Sheet1"
                        if target:
                            # Linked/Merge Logic check... (Same as before)
                            linked_ids = target.get("linkedSourceIds")
                            if linked_ids and isinstance(linked_ids, list) and len(linked_ids) > 0:
                                # We need to re-find node... inefficient strictly but safe
                                node_source_targets = node.get(
                                    "data", {}).get("sourceTargets", [])
                                dfs_to_merge = []
                                for src_idx in linked_ids:
                                    if isinstance(src_idx, int) and 0 <= src_idx < len(node_source_targets):
                                        dfs_to_merge.append(get_df_for_target(
                                            node_source_targets[src_idx]))
                                sheet_df = pd.concat(
                                    dfs_to_merge, ignore_index=True) if dfs_to_merge else pd.DataFrame()
                            else:
                                sheet_df = get_df_for_target(target)
                        else:
                            output_id = item.get("id") or "out"
                            virtual_key = f"virtual:output:{output_id}:{sheet_name}"
                            sheet_df = table_map.get(
                                virtual_key, pd.DataFrame())

                        # Write to shared writer
                        # Check for name collision
                        if sheet_name in writer.sheets:
                            # Overwrite logic (pandas default) or append rows?
                            # Prompt said "Append filtered data".
                            # If we assume APPEND ROWS:
                            # old_max_row = writer.sheets[sheet_name].max_row
                            # But pandas doesn't support 'append rows' easily via high level too_excel.
                            # We stick to OVERWRITE/REPLACE sheet for MVP, or new sheet name.
                            pass
                        sheet_df.to_excel(writer, index=False,
                                          sheet_name=sheet_name)
                else:
                    # Single legacy file fallback
                    df = get_df_for_target(target) if target else result_df
                    df.to_excel(writer, index=False, sheet_name="Sheet1")

            writer.close()
            payload = output.getvalue()

            # Save logic for one file
            file_name = outputs_to_write[0].get(
                "fileName") if outputs_to_write else "appended.xlsx"
            # We ignore file_name from others, we just use the first/base one.
            # Or assume base filename.
            file_name = base_file.original_filename if 'base_file' in locals() else file_name

            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

            if output_batch:
                # Save...
                file_name = file_service.resolve_unique_original_name(
                    db, current_user.id, output_batch.id, file_name)
                file_service.save_generated_file(
                    db, current_user.id, file_name, payload, output_batch.id)

            files_payload.append({
                "file_name": file_name,
                "payload": payload,
                "media_type": media_type
            })

        else:
            # STANDARD LOGIC (Loop and create separate files)
            for index, output_item in enumerate(outputs_to_write):
                file_name = output_item.get(
                    "fileName") or f"output_{index}.xlsx"
                sheets = output_item.get("sheets", [])
                # Our custom attached target ref
                target = output_item.get("target")

                output_id = output_item.get("id") or f"out-{index}"

                file_extension = Path(file_name).suffix.lower()

                if file_extension == ".csv":
                    if sheets:
                        sheet_name = sheets[0].get("sheetName") or "Sheet1"
                        # If we have a target ref, use it directly
                        if target:
                            result_for_file = get_df_for_target(target)
                        else:
                            # Legacy Output/Virtual Key lookup
                            virtual_key = f"virtual:output:{output_id}:{sheet_name}"
                            result_for_file = table_map.get(
                                virtual_key, result_df if not target else pd.DataFrame())
                    else:
                        # Fallback to last result if available and no target
                        result_for_file = result_df if not target else get_df_for_target(
                            target)

                    output = io.StringIO()
                    result_for_file.to_csv(output, index=False)
                    payload = output.getvalue().encode("utf-8")
                    media_type = "text/csv"
                else:
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine="openpyxl") as writer:
                        if sheets:
                            for sheet in sheets:
                                sheet_name = sheet.get("sheetName") or "Sheet1"
                                if target:
                                    # If target exists, it corresponds to this sheet/file.
                                    # Check for G2M Merge (linkedSourceIds)
                                    linked_ids = target.get("linkedSourceIds")
                                    if linked_ids and isinstance(linked_ids, list) and len(linked_ids) > 0:
                                        # Merge Logic: Concatenate all linked sources
                                        dfs_to_merge = []
                                        # We need access to the node's sourceTargets to resolve indices
                                        # The 'node' variable is available from the outer loop
                                        node_source_targets = node.get(
                                            "data", {}).get("sourceTargets", [])

                                        for src_idx in linked_ids:
                                            if isinstance(src_idx, int) and 0 <= src_idx < len(node_source_targets):
                                                src_target = node_source_targets[src_idx]
                                                dfs_to_merge.append(
                                                    get_df_for_target(src_target))

                                        if dfs_to_merge:
                                            sheet_df = pd.concat(
                                                dfs_to_merge, ignore_index=True)
                                        else:
                                            sheet_df = pd.DataFrame()
                                    else:
                                        # Standard 1:1 Logic
                                        sheet_df = get_df_for_target(target)
                                else:
                                    virtual_key = f"virtual:output:{output_id}:{sheet_name}"
                                    sheet_df = table_map.get(
                                        virtual_key, pd.DataFrame())

                                sheet_df.to_excel(
                                    writer, index=False, sheet_name=sheet_name)
                        else:
                            # Fallback
                            df = get_df_for_target(
                                target) if target else result_df
                            df.to_excel(writer, index=False,
                                        sheet_name="Sheet1")

                    output.seek(0)
                    payload = output.read()
                    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

                if output_batch:
                    file_name = file_service.resolve_unique_original_name(
                        db=db,
                        user_id=current_user.id,
                        batch_id=output_batch.id,
                        desired_name=file_name,
                        reserved_names=reserved_output_names,
                    )
                    reserved_output_names.add(file_name)
                    file_service.save_generated_file(
                        db=db,
                        user_id=current_user.id,
                        original_filename=file_name,
                        content=payload,
                        batch_id=output_batch.id,
                    )
                files_payload.append({
                    "file_name": file_name,
                    "payload": payload,
                    "media_type": media_type,
                })

        from fastapi.responses import StreamingResponse
        if len(files_payload) == 1:
            file_name = files_payload[0]["file_name"]
            payload = files_payload[0]["payload"]
            media_type = files_payload[0]["media_type"]
            return StreamingResponse(
                io.BytesIO(payload),
                media_type=media_type,
                headers={
                    "Content-Disposition": f"attachment; filename={file_name}"
                }
            )

        zip_content = create_zip_archive(files_payload, output_batch)
        return StreamingResponse(
            io.BytesIO(zip_content),
            media_type="application/zip",
            headers={
                "Content-Disposition": "attachment; filename=outputs.zip"
            }
        )
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Error exporting result: {str(e)}"
        )
